import os
import io
from datetime import datetime, date, timedelta

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, login_user, login_required, logout_user,
    current_user, UserMixin
)
from sqlalchemy import func, text, inspect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.styles import Font

# ------------------ App config ------------------
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("APP_SECRET_KEY", "dev-secret")

# Usa DATABASE_URL su Render se presente, altrimenti sqlite locale
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///data.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

BASE_DIR = os.path.dirname(__file__)
BASE_STATIC = os.path.join(BASE_DIR, "static")
UPLOAD_ROOT = os.environ.get("UPLOAD_ROOT", os.path.join(BASE_STATIC, "uploads"))
os.makedirs(UPLOAD_ROOT, exist_ok=True)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif", "pdf"}
def allowed_file(name: str) -> bool:
    return "." in name and name.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"

# ------------------ Models ------------------
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), default="capo")  # "admin" o "capo"
    full_name = db.Column(db.String(120))

    def set_password(self, p):
        self.password_hash = generate_password_hash(p)

    def check_password(self, p):
        return check_password_hash(self.password_hash, p)

class Client(db.Model):
    __tablename__ = "clients"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    note = db.Column(db.String(255))  # auto-patched se mancante

class Site(db.Model):
    __tablename__ = "sites"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    address = db.Column(db.String(255))  # auto-patched se mancante
    is_confirmed = db.Column(db.Integer, default=0)  # 0/1; auto-patched se mancante
    client_id = db.Column(db.Integer, db.ForeignKey("clients.id"), nullable=False)
    client = db.relationship("Client")

class ActivityCatalog(db.Model):
    __tablename__ = "activity_catalog"
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(40), nullable=False, unique=True)
    description = db.Column(db.String(255), nullable=False)
    unit = db.Column(db.String(20), nullable=False)
    unit_price = db.Column(db.Float, nullable=False, default=0.0)  # admin può impostare

class ClientActivity(db.Model):
    __tablename__ = "client_activity"
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey("sites.id"), nullable=False)
    activity_id = db.Column(db.Integer, db.ForeignKey("activity_catalog.id"), nullable=False)
    initial_qty = db.Column(db.Float, default=0.0)
    site = db.relationship("Site")
    activity = db.relationship("ActivityCatalog")

class Vehicle(db.Model):
    __tablename__ = "vehicles"
    id = db.Column(db.Integer, primary_key=True)
    plate = db.Column(db.String(20), unique=True, nullable=False)
    description = db.Column(db.String(120))
    site_id = db.Column(db.Integer, db.ForeignKey("sites.id"))
    site = db.relationship("Site")

class Equipment(db.Model):
    __tablename__ = "equipment"
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(40), unique=True, nullable=False)
    description = db.Column(db.String(255), nullable=False)
    qty = db.Column(db.Integer, default=1)
    status = db.Column(db.String(20), default="disponibile")
    site_id = db.Column(db.Integer, db.ForeignKey("sites.id"))
    site = db.relationship("Site")

class Assignment(db.Model):
    __tablename__ = "assignments"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    site_id = db.Column(db.Integer, db.ForeignKey("sites.id"), nullable=False)
    start_date = db.Column(db.Date)  # auto-patched se mancante
    end_date = db.Column(db.Date)    # auto-patched se mancante
    user = db.relationship("User")
    site = db.relationship("Site")

class ActivityEntry(db.Model):
    """Rilevazione giornaliera quantità lavorata (capo)"""
    __tablename__ = "activity_entries"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    site_id = db.Column(db.Integer, db.ForeignKey("sites.id"), nullable=False)
    client_activity_id = db.Column(db.Integer, db.ForeignKey("client_activity.id"), nullable=False)
    work_date = db.Column(db.Date, nullable=False, default=date.today)
    qty = db.Column(db.Float, nullable=False, default=0.0)
    note = db.Column(db.String(255))
    photo_path = db.Column(db.String(255))  # capo può caricare foto (auto-patched se mancante)
    user = db.relationship("User")
    site = db.relationship("Site")
    client_activity = db.relationship("ClientActivity")

class ExtraActivity(db.Model):
    """
    Attività Extra (non a catalogo), inserita dal capo con foto.
    L'admin può 'quotare' (unit_price) prima di includerla nei report.
    """
    __tablename__ = "extra_activities"
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey("sites.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    work_date = db.Column(db.Date, nullable=False, default=date.today)
    description = db.Column(db.String(255), nullable=False)
    unit = db.Column(db.String(20), nullable=False, default="u")
    qty = db.Column(db.Float, nullable=False, default=1.0)
    photo_path = db.Column(db.String(255))
    unit_price = db.Column(db.Float)  # se None => non ancora quotata
    approved = db.Column(db.Boolean, default=False)

    site = db.relationship("Site")
    user = db.relationship("User")

class SiteExpense(db.Model):
    __tablename__ = "site_expenses"
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey("sites.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    exp_type = db.Column(db.String(120), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_type = db.Column(db.String(40))
    exp_date = db.Column(db.Date, default=date.today, nullable=False)
    receipt_path = db.Column(db.String(255))
    site = db.relationship("Site")
    user = db.relationship("User")

class VehicleExpense(db.Model):
    __tablename__ = "vehicle_expenses"
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey("vehicles.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    exp_type = db.Column(db.String(40), nullable=False)  # CARBURANTE/PEDAGGIO/PARCHEGGIO
    amount = db.Column(db.Float, nullable=False)
    exp_date = db.Column(db.Date, default=date.today, nullable=False)
    receipt_path = db.Column(db.String(255))
    vehicle = db.relationship("Vehicle")
    user = db.relationship("User")

class Alert(db.Model):
    __tablename__ = "alerts"
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    message = db.Column(db.String(255), nullable=False)
    is_read = db.Column(db.Boolean, default=False)

# ------------------ Login ------------------
@login_manager.user_loader
def load_user(uid):
    return User.query.get(int(uid))

# ------------------ Helpers ------------------
def is_admin():
    return current_user.is_authenticated and current_user.role == "admin"

def save_upload(file_storage, prefix):
    if not file_storage or not file_storage.filename:
        return None
    fname = secure_filename(file_storage.filename)
    if not allowed_file(fname):
        return None
    out = f"{prefix}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{fname}"
    path = os.path.join(UPLOAD_ROOT, out)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    file_storage.save(path)
    # Se UPLOAD_ROOT sta in /static, torna un path relativo servibile
    if UPLOAD_ROOT.startswith(BASE_STATIC):
        return f"uploads/{out}"
    return path

# ------------------ Auto-migrazioni SQLite (compatibile SQLAlchemy 2.x) ------------------
def has_column(table: str, col: str) -> bool:
    insp = inspect(db.engine)
    try:
        return col in [c["name"] for c in insp.get_columns(table)]
    except Exception:
        return False

def ensure_sqlite_columns():
    with db.engine.connect() as conn:
        # clients.note
        if has_column("clients", "id") and not has_column("clients", "note"):
            conn.execute(text("ALTER TABLE clients ADD COLUMN note TEXT"))
        # sites.address
        if has_column("sites", "id") and not has_column("sites", "address"):
            conn.execute(text("ALTER TABLE sites ADD COLUMN address TEXT"))
        # sites.is_confirmed
        if has_column("sites", "id") and not has_column("sites", "is_confirmed"):
            conn.execute(text("ALTER TABLE sites ADD COLUMN is_confirmed INTEGER DEFAULT 0"))
        # assignments.start_date / end_date
        if has_column("assignments", "id") and not has_column("assignments", "start_date"):
            conn.execute(text("ALTER TABLE assignments ADD COLUMN start_date DATE"))
        if has_column("assignments", "id") and not has_column("assignments", "end_date"):
            conn.execute(text("ALTER TABLE assignments ADD COLUMN end_date DATE"))
        # activity_entries.photo_path
        if has_column("activity_entries", "id") and not has_column("activity_entries", "photo_path"):
            conn.execute(text("ALTER TABLE activity_entries ADD COLUMN photo_path TEXT"))
        try:
            conn.commit()
        except Exception:
            pass

# ------------------ Demo seed ------------------
def ensure_demo():
    db.create_all()
    if not User.query.filter_by(username="admin").first():
        admin = User(username="admin", role="admin", full_name="Amministratore")
        admin.set_password("admin")
        db.session.add(admin)
    if not User.query.filter_by(username="capo").first():
        capo = User(username="capo", role="capo", full_name="Mario Rossi")
        capo.set_password("capo")
        db.session.add(capo)
    if not Client.query.first():
        c = Client(name="Cliente Alpha")
        db.session.add(c)
        s1 = Site(name="Cantiere Nord", client=c, address="Via Roma 1", is_confirmed=1)
        s2 = Site(name="Cantiere Sud", client=c, address="Via Milano 2", is_confirmed=0)
        db.session.add_all([s1, s2])
    if not ActivityCatalog.query.first():
        db.session.add_all([
            ActivityCatalog(code="SCAVO", description="Scavo generale", unit="m3", unit_price=25),
            ActivityCatalog(code="MUR", description="Muratura", unit="mq", unit_price=45),
        ])
    db.session.commit()

with app.app_context():
    ensure_demo()
    ensure_sqlite_columns()

# ------------------ Routes: base/auth ------------------
@app.route("/", endpoint="index")
def index():
    # Se i template aspettano index, li portiamo al login
    return redirect(url_for("login"))

@app.route("/admin", endpoint="admin_home")
@login_required
def admin_home():
    if not is_admin():
        return redirect(url_for("dashboard"))
    unread = Alert.query.filter_by(is_read=False).count()
    return render_template("dashboard_admin.html", unread=unread)

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","")
        u = User.query.filter_by(username=username).first()
        if u and u.check_password(password):
            login_user(u)
            flash("Accesso eseguito", "success")
            return redirect(url_for("dashboard"))
        flash("Credenziali errate", "danger")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logout effettuato", "info")
    return redirect(url_for("index"))

@app.route("/dashboard")
@login_required
def dashboard():
    if is_admin():
        unread = Alert.query.filter_by(is_read=False).count()
        return render_template("dashboard_admin.html", unread=unread)
    # capo
    my_sites = (
        db.session.query(Site)
        .join(Assignment, Assignment.site_id == Site.id)
        .filter(Assignment.user_id == current_user.id)
        .all()
    )
    return render_template("dashboard_capo.html", sites=my_sites)

# ------------------ Admin: Clienti & Cantieri ------------------
@app.route("/admin/clients-sites", methods=["GET","POST"])
@login_required
def clients_sites():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        kind = request.form.get("kind")
        name = request.form.get("name","").strip()
        if not name:
            flash("Nome obbligatorio", "warning"); return redirect(url_for("clients_sites"))
        if kind == "client":
            note = request.form.get("note","").strip() or None
            db.session.add(Client(name=name, note=note))
        elif kind == "site":
            client_id = request.form.get("client_id", type=int)
            address = request.form.get("address","").strip() or None
            is_confirmed = 1 if request.form.get("is_confirmed") == "1" else 0
            if not client_id:
                flash("Cliente obbligatorio per il cantiere", "warning"); return redirect(url_for("clients_sites"))
            db.session.add(Site(name=name, client_id=client_id, address=address, is_confirmed=is_confirmed))
        db.session.commit()
        flash("Salvato", "success")
        return redirect(url_for("clients_sites"))
    return render_template("clients_sites.html",
                           clients=Client.query.order_by(Client.name).all(),
                           sites=Site.query.order_by(Site.name).all())

@app.route("/admin/clients-sites/<string:kind>/<int:oid>/delete", methods=["POST"])
@login_required
def delete_clients_sites(kind, oid):
    if not is_admin(): return redirect(url_for("dashboard"))
    model = Client if kind=="client" else Site
    obj = model.query.get_or_404(oid)
    db.session.delete(obj); db.session.commit()
    flash("Eliminato", "success")
    return redirect(url_for("clients_sites"))

# ------------------ Admin: Catalogo (prezzo attività) & Assegnazioni ------------------
@app.route("/admin/catalogo", methods=["GET","POST"])
@login_required
def catalogo():
    if not is_admin():
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        code = request.form.get("code","").strip().upper()
        description = request.form.get("description","").strip()
        unit = request.form.get("unit","").strip()
        unit_price = request.form.get("unit_price", type=float) or 0.0

        if not code or not description or not unit:
            flash("Campi mancanti", "warning")
        else:
            if ActivityCatalog.query.filter_by(code=code).first():
                flash("Codice già presente", "danger")
            else:
                db.session.add(ActivityCatalog(
                    code=code,
                    description=description,
                    unit=unit,
                    unit_price=unit_price
                ))
                db.session.commit()
                flash("Attività inserita", "success")
        return redirect(url_for("catalogo"))

    # aggiungo anche i cantieri per il template
    sites = Site.query.join(Client).order_by(Site.name).all()

    return render_template(
        "catalogo.html",
        items=ActivityCatalog.query.order_by(ActivityCatalog.code).all(),
        sites=sites
    )

@app.route("/admin/catalogo/<int:aid>/prezzo", methods=["POST"])
@login_required
def update_prezzo_catalogo(aid):
    """Admin aggiorna il valore economico (prezzo unitario) delle attività a catalogo"""
    if not is_admin(): return redirect(url_for("dashboard"))
    a = ActivityCatalog.query.get_or_404(aid)
    unit_price = request.form.get("unit_price", type=float)
    if unit_price is None:
        flash("Prezzo non valido", "warning")
    else:
        a.unit_price = float(unit_price)
        db.session.commit()
        flash("Prezzo aggiornato", "success")
    return redirect(url_for("catalogo"))

@app.route("/admin/catalogo/<int:aid>/delete", methods=["POST"])
@login_required
def delete_catalogo(aid):
    if not is_admin(): return redirect(url_for("dashboard"))
    a = ActivityCatalog.query.get_or_404(aid)
    db.session.delete(a); db.session.commit()
    flash("Eliminato", "success")
    return redirect(url_for("catalogo"))

@app.route("/admin/assign-activity", methods=["POST"])
@login_required
def assign_activity():
    if not is_admin(): return redirect(url_for("dashboard"))
    site_id = request.form.get("site_id", type=int)
    activity_id = request.form.get("activity_id", type=int)
    initial_qty = request.form.get("initial_qty", type=float) or 0.0
    if not site_id or not activity_id:
        flash("Seleziona cantiere e attività", "warning")
    else:
        db.session.add(ClientActivity(site_id=site_id, activity_id=activity_id, initial_qty=initial_qty))
        db.session.commit()
        flash("Assegnazione creata", "success")
    return redirect(url_for("catalogo"))

# ------------------ Admin: Utenti & Assegnazioni cantieri ------------------
@app.route("/admin/users", methods=["GET","POST"])
@login_required
def users():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","")
        role = request.form.get("role","capo")
        full_name = request.form.get("full_name","").strip()
        if not username or not password:
            flash("Username e password obbligatori", "warning")
        else:
            if User.query.filter_by(username=username).first():
                flash("Username già in uso", "danger")
            else:
                u = User(username=username, role=role, full_name=full_name)
                u.set_password(password)
                db.session.add(u); db.session.commit()
                flash("Utente creato", "success")
        return redirect(url_for("users"))
    return render_template("users.html",
                           users=User.query.order_by(User.username).all(),
                           sites=Site.query.order_by(Site.name).all(),
                           assignments=Assignment.query.all())

@app.route("/admin/users/<int:uid>/update", methods=["POST"])
@login_required
def update_user(uid):
    if not is_admin(): return redirect(url_for("dashboard"))
    u = User.query.get_or_404(uid)
    username = (request.form.get("username") or u.username).strip()
    full_name = (request.form.get("full_name") or u.full_name or "").strip()
    role = (request.form.get("role") or u.role or "capo").strip()
    password = request.form.get("password")

    if u.id == current_user.id and role != "admin":
        flash("Non puoi rimuovere il ruolo admin da te stesso.", "warning")
        return redirect(url_for("users"))

    if User.query.filter(User.username == username, User.id != u.id).first():
        flash("Username già in uso da un altro utente.", "danger")
        return redirect(url_for("users"))

    u.username, u.full_name, u.role = username, full_name, role
    if password and password.strip():
        u.set_password(password.strip())
    try:
        db.session.commit()
        flash("Utente aggiornato.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Errore aggiornamento: {e}", "danger")
    return redirect(url_for("users"))

@app.route("/admin/users/<int:uid>/delete", methods=["POST"])
@login_required
def delete_user(uid):
    if not is_admin(): return redirect(url_for("dashboard"))
    if current_user.id == uid:
        flash("Non puoi eliminare te stesso", "warning")
        return redirect(url_for("users"))
    u = User.query.get_or_404(uid)
    db.session.delete(u); db.session.commit()
    flash("Utente eliminato", "success")
    return redirect(url_for("users"))

@app.route("/admin/assign-site", methods=["POST"])
@login_required
def assign_site():
    if not is_admin(): return redirect(url_for("dashboard"))
    user_id = request.form.get("user_id", type=int)
    site_id = request.form.get("site_id", type=int)
    start_date = request.form.get("start_date")
    end_date = request.form.get("end_date")
    if user_id and site_id:
        if not Assignment.query.filter_by(user_id=user_id, site_id=site_id).first():
            a = Assignment(user_id=user_id, site_id=site_id)
            if start_date:
                a.start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            if end_date:
                a.end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            db.session.add(a); db.session.commit()
            flash("Cantiere assegnato", "success")
    return redirect(url_for("users"))

@app.route("/admin/assign-site/<int:aid>/delete", methods=["POST"])
@login_required
def delete_assignment(aid):
    if not is_admin(): return redirect(url_for("dashboard"))
    a = Assignment.query.get_or_404(aid)
    db.session.delete(a); db.session.commit()
    flash("Assegnazione rimossa", "success")
    return redirect(url_for("users"))

# ------------------ Admin: Veicoli & Attrezzature ------------------
@app.route("/admin/vehicles", methods=["GET","POST"])
@login_required
def vehicles():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        plate = request.form.get("plate","").upper().strip()
        description = request.form.get("description","").strip()
        site_id = request.form.get("site_id", type=int)
        if not plate:
            flash("Targa obbligatoria", "warning")
        else:
            db.session.add(Vehicle(plate=plate, description=description, site_id=site_id))
            db.session.commit()
            flash("Veicolo inserito", "success")
        return redirect(url_for("vehicles"))
    return render_template("vehicles.html",
                           vehicles=Vehicle.query.order_by(Vehicle.plate).all(),
                           sites=Site.query.order_by(Site.name).all())

@app.route("/admin/vehicles/<int:vid>/delete", methods=["POST"])
@login_required
def delete_vehicle(vid):
    if not is_admin(): return redirect(url_for("dashboard"))
    v = Vehicle.query.get_or_404(vid)
    db.session.delete(v); db.session.commit()
    flash("Veicolo eliminato", "success")
    return redirect(url_for("vehicles"))

@app.route("/admin/equipment", methods=["GET","POST"])
@login_required
def equipment():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        code = request.form.get("code","").strip().upper()
        description = request.form.get("description","").strip()
        qty = request.form.get("qty", type=int) or 1
        site_id = request.form.get("site_id", type=int)
        status = request.form.get("status","disponibile")
        if not code or not description:
            flash("Codice e descrizione obbligatori", "warning")
        else:
            db.session.add(Equipment(code=code, description=description, qty=qty, site_id=site_id, status=status))
            db.session.commit()
            flash("Attrezzatura inserita", "success")
        return redirect(url_for("equipment"))

    alerts = Alert.query.order_by(Alert.created_at.desc()).limit(15).all()
    return render_template("equipment.html",
                           equipments=Equipment.query.order_by(Equipment.code).all(),
                           sites=Site.query.order_by(Site.name).all(),
                           alerts=alerts)

@app.route("/admin/equipment/<int:eid>/status", methods=["POST"])
@login_required
def equipment_status(eid):
    if not is_admin(): return redirect(url_for("dashboard"))
    e = Equipment.query.get_or_404(eid)
    new_status = request.form.get("status","disponibile")
    e.status = new_status
    db.session.commit()
    flash("Stato aggiornato", "success")
    return redirect(url_for("equipment"))

@app.route("/admin/equipment/<int:eid>/delete", methods=["POST"], endpoint="equipment_delete")
@login_required
def delete_equipment(eid):
    if not is_admin(): return redirect(url_for("dashboard"))
    e = Equipment.query.get_or_404(eid)
    db.session.delete(e); db.session.commit()
    flash("Attrezzatura eliminata", "success")
    return redirect(url_for("equipment"))

# ------------------ Capo: Attività (con foto) ------------------
@app.route("/capo/attivita", methods=["GET","POST"])
@login_required
def capo_attivita():
    my_assignments = Assignment.query.all() if is_admin() else Assignment.query.filter_by(user_id=current_user.id).all()

    if request.method == "POST":
        work_date = request.form.get("work_date") or date.today().isoformat()
        site_id = request.form.get("site_id", type=int)
        client_activity_id = request.form.get("client_activity_id", type=int)
        qty = request.form.get("qty", type=float) or 0.0
        note = request.form.get("note","").strip() or None
        photo = save_upload(request.files.get("photo"), "ACT")
        if not site_id or not client_activity_id:
            flash("Seleziona cantiere e attività", "warning")
        else:
            entry = ActivityEntry(
                user_id=current_user.id,
                site_id=site_id,
                client_activity_id=client_activity_id,
                work_date=datetime.strptime(work_date,"%Y-%m-%d").date(),
                qty=qty,
                note=note,
                photo_path=photo
            )
            db.session.add(entry); db.session.commit()
            flash("Rilevazione salvata", "success")
        return redirect(url_for("capo_attivita"))

    site_ids = [a.site_id for a in my_assignments]
    assigned_ca = ClientActivity.query.filter(ClientActivity.site_id.in_(site_ids)).all() if site_ids else []
    entries = ActivityEntry.query.order_by(ActivityEntry.work_date.desc(), ActivityEntry.id.desc())
    if not is_admin():
        entries = entries.filter(ActivityEntry.user_id == current_user.id)
    entries = entries.limit(200).all()
    return render_template("capo_attivita.html", assignments=my_assignments, assigned_ca=assigned_ca, entries=entries)

@app.route("/capo/attivita/<int:eid>/delete", methods=["POST"])
@login_required
def delete_entry(eid):
    e = ActivityEntry.query.get_or_404(eid)
    if not is_admin() and e.user_id != current_user.id:
        flash("Operazione non consentita", "danger"); return redirect(url_for("capo_attivita"))
    db.session.delete(e); db.session.commit()
    flash("Eliminato", "success")
    return redirect(url_for("capo_attivita"))

# ------------------ Capo: Attività EXTRA (con foto) ------------------
@app.route("/capo/attivita-extra", methods=["GET","POST"])
@login_required
def capo_attivita_extra():
    assignments = Assignment.query.filter_by(user_id=current_user.id).all() if not is_admin() else Assignment.query.all()
    if request.method == "POST":
        site_id = request.form.get("site_id", type=int)
        work_date = request.form.get("work_date") or date.today().isoformat()
        description = request.form.get("description","").strip()
        unit = request.form.get("unit","").strip() or "u"
        qty = request.form.get("qty", type=float) or 1.0
        photo = save_upload(request.files.get("photo"), "EXTRA")
        if not site_id or not description:
            flash("Compila i campi obbligatori", "warning")
        else:
            row = ExtraActivity(
                site_id=site_id,
                user_id=current_user.id,
                work_date=datetime.strptime(work_date,"%Y-%m-%d").date(),
                description=description,
                unit=unit,
                qty=qty,
                photo_path=photo
            )
            db.session.add(row); db.session.commit()
            flash("Attività extra inviata (in attesa di quotazione)", "success")
        return redirect(url_for("capo_attivita_extra"))

    q = ExtraActivity.query.order_by(ExtraActivity.work_date.desc(), ExtraActivity.id.desc())
    if not is_admin():
        q = q.filter(ExtraActivity.user_id == current_user.id)
    rows = q.limit(200).all()
    return render_template("capo_attivita_extra.html", assignments=assignments, rows=rows)

# ------------------ Admin: Gestione & quotazione Attività EXTRA ------------------
@app.route("/admin/attivita-extra", methods=["GET","POST"])
@login_required
def admin_attivita_extra():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        eid = request.form.get("extra_id", type=int)
        unit_price = request.form.get("unit_price", type=float)
        approved = True if request.form.get("approved") == "1" else False
        row = ExtraActivity.query.get_or_404(eid)
        row.unit_price = unit_price
        row.approved = approved
        db.session.commit()
        flash("Quotazione aggiornata", "success")
        return redirect(url_for("admin_attivita_extra"))

    sites = Site.query.order_by(Site.name.asc()).all()
    rows = ExtraActivity.query.order_by(ExtraActivity.work_date.desc()).all()
    return render_template("admin_attivita_extra.html", sites=sites, rows=rows)

# ------------------ Spese cantiere (capo) ------------------
@app.route("/capo/spese", methods=["GET","POST"])
@login_required
def capo_spese():
    assignments = Assignment.query.filter_by(user_id=current_user.id).all() if not is_admin() else Assignment.query.all()
    if request.method == "POST":
        site_id = request.form.get("site_id", type=int)
        exp_type = request.form.get("exp_type","").strip()
        amount = request.form.get("amount", type=float) or 0.0
        payment_type = request.form.get("payment_type","").strip() or None
        exp_date = request.form.get("exp_date") or date.today().isoformat()
        receipt = save_upload(request.files.get("receipt"), "SITEEXP")
        if not site_id or not exp_type or not amount:
            flash("Campi obbligatori mancanti", "warning")
        else:
            db.session.add(SiteExpense(
                site_id=site_id, user_id=current_user.id, exp_type=exp_type,
                amount=amount, payment_type=payment_type,
                exp_date=datetime.strptime(exp_date,"%Y-%m-%d").date(),
                receipt_path=receipt
            ))
            db.session.commit()
            flash("Spesa registrata", "success")
        return redirect(url_for("capo_spese"))
    q = SiteExpense.query.order_by(SiteExpense.exp_date.desc(), SiteExpense.id.desc())
    if not is_admin():
        q = q.filter(SiteExpense.user_id == current_user.id)
    rows = q.limit(200).all()
    return render_template("capo_spese.html", assignments=assignments, rows=rows)

# ------------------ Spese veicoli (capo) ------------------
@app.route("/capo/spese-veicoli", methods=["GET","POST"])
@login_required
def capo_spese_veicoli():
    vehicles = Vehicle.query.order_by(Vehicle.plate).all()
    if request.method == "POST":
        vehicle_id = request.form.get("vehicle_id", type=int)
        exp_type = request.form.get("exp_type","").strip().upper()
        amount = request.form.get("amount", type=float) or 0.0
        exp_date = request.form.get("exp_date") or date.today().isoformat()
        receipt = save_upload(request.files.get("receipt"), "VEH")
        if not vehicle_id or not exp_type or not amount:
            flash("Campi obbligatori mancanti", "warning")
        else:
            db.session.add(VehicleExpense(
                vehicle_id=vehicle_id, user_id=current_user.id,
                exp_type=exp_type, amount=amount,
                exp_date=datetime.strptime(exp_date,"%Y-%m-%d").date(),
                receipt_path=receipt
            ))
            db.session.commit()
            flash("Spesa veicolo registrata", "success")
        return redirect(url_for("capo_spese_veicoli"))
    q = VehicleExpense.query.order_by(VehicleExpense.exp_date.desc(), VehicleExpense.id.desc())
    if not is_admin():
        q = q.filter(VehicleExpense.user_id == current_user.id)
    rows = q.limit(200).all()
    return render_template("capo_spese_veicoli.html", vehicles=vehicles, rows=rows)

# ------------------ Capo: segnalazione attrezzature ------------------
@app.route("/capo/equipment/<int:eid>/status", methods=["POST"])
@login_required
def capo_equipment_status(eid):
    e = Equipment.query.get_or_404(eid)
    new_status = request.form.get("status","manutenzione")
    if new_status not in {"rotto", "manutenzione"}:
        new_status = "manutenzione"
    e.status = new_status
    db.session.add(Alert(message=f"Attrezzatura {e.code} segnalata '{new_status}' dal capo {current_user.username}"))
    db.session.commit()
    flash("Segnalazione inviata", "success")
    return redirect(url_for("dashboard"))

# ------------------ Report ed export ------------------
@app.route("/admin/report", methods=["GET"])
@login_required
def admin_report():
    if not is_admin(): return redirect(url_for("dashboard"))
    today = date.today()
    start = datetime(today.year, today.month, 1).date()

    # Valore attività (catalogo): somma qty * unit_price
    acts = db.session.query(
        func.sum(ActivityEntry.qty * ActivityCatalog.unit_price)
    ).join(ClientActivity, ClientActivity.id == ActivityEntry.client_activity_id)\
     .join(ActivityCatalog, ActivityCatalog.id == ClientActivity.activity_id)\
     .filter(ActivityEntry.work_date >= start).scalar() or 0.0

    # Valore attività EXTRA approvate e quotate
    extra_val = db.session.query(
        func.sum(ExtraActivity.qty * ExtraActivity.unit_price)
    ).filter(
        ExtraActivity.work_date >= start,
        ExtraActivity.approved == True,
        ExtraActivity.unit_price.isnot(None)
    ).scalar() or 0.0

    veh_cost = db.session.query(func.sum(VehicleExpense.amount)).scalar() or 0.0
    site_cost = db.session.query(func.sum(SiteExpense.amount)).scalar() or 0.0

    total_val = float(acts) + float(extra_val)

    start_default = (date.today() - timedelta(days=7)).isoformat()
    end_default = date.today().isoformat()

    return render_template(
        "admin_report.html",
        total_val=total_val,
        extra_val=extra_val,
        veh_cost=veh_cost,
        site_cost=site_cost,
        start_default=start_default,
        end_default=end_default
    )

@app.route("/admin/export/attivita")
@login_required
def export_attivita():
    if not is_admin(): return redirect(url_for("dashboard"))
    start = request.args.get("start") or (date.today() - timedelta(days=7)).isoformat()
    end = request.args.get("end") or date.today().isoformat()
    sdate = datetime.strptime(start,"%Y-%m-%d").date()
    edate = datetime.strptime(end,"%Y-%m-%d").date()

    rows = db.session.query(
        ActivityEntry.work_date, Site.name.label("site_name"), Client.name.label("client_name"),
        ActivityCatalog.code, ActivityCatalog.description, ActivityCatalog.unit,
        ActivityEntry.qty, ActivityCatalog.unit_price,
        (ActivityEntry.qty * ActivityCatalog.unit_price).label("value")
    ).join(Site, Site.id == ActivityEntry.site_id)\
     .join(Client, Client.id == Site.client_id)\
     .join(ClientActivity, ClientActivity.id == ActivityEntry.client_activity_id)\
     .join(ActivityCatalog, ActivityCatalog.id == ClientActivity.activity_id)\
     .filter(ActivityEntry.work_date >= sdate, ActivityEntry.work_date <= edate)\
     .order_by(ActivityEntry.work_date).all()

    wb = Workbook(); ws = wb.active; ws.title = "Attività"
    ws.append(["Data","Cantiere","Cliente","Codice","Descrizione","UM","Q.tà","Prezzo Unit.","Valore (€)"])
    for c in range(1,10): ws.cell(row=1, column=c).font = Font(bold=True)
    total = 0.0
    for r in rows:
        total += float(r.value or 0.0)
        ws.append([
            r.work_date.strftime("%Y-%m-%d"),
            r.site_name, r.client_name, r.code, r.description, r.unit,
            float(r.qty or 0), float(r.unit_price or 0), float(r.value or 0)
        ])

    # Extra approvate e quotate nel range
    extra = ExtraActivity.query.filter(
        ExtraActivity.work_date >= sdate, ExtraActivity.work_date <= edate,
        ExtraActivity.approved == True, ExtraActivity.unit_price.isnot(None)
    ).order_by(ExtraActivity.work_date).all()

    ws.append([])
    ws.append(["ATTIVITÀ EXTRA APPROVATE"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Data","Cantiere","Descrizione","UM","Q.tà","Prezzo Unit.","Valore (€)"])
    for c in range(1,8): ws.cell(row=ws.max_row, column=c).font = Font(bold=True)

    extra_tot = 0.0
    for e in extra:
        val = float((e.qty or 0.0) * (e.unit_price or 0.0))
        extra_tot += val
        ws.append([
            e.work_date.strftime("%Y-%m-%d"),
            e.site.name if e.site else "",
            e.description, e.unit, float(e.qty or 0.0),
            float(e.unit_price or 0.0), val
        ])

    ws.append([])
    ws.append(["","", "Totale Attività Catalogo", "", "", "", total])
    ws.append(["","", "Totale Attività Extra", "", "", "", extra_tot])
    ws.append(["","", "TOTALE COMPLESSIVO", "", "", "", total + extra_tot])
    for i in range(3,8,3):
        ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
        ws.cell(row=ws.max_row-1, column=3).font = Font(bold=True)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"attivita_{sdate.strftime('%Y%m%d')}_{edate.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/export/spese-veicoli")
@login_required
def export_spese_veicoli():
    if not is_admin(): return redirect(url_for("dashboard"))
    start = request.args.get("start") or (date.today() - timedelta(days=7)).isoformat()
    end = request.args.get("end") or date.today().isoformat()
    sdate = datetime.strptime(start,"%Y-%m-%d").date()
    edate = datetime.strptime(end,"%Y-%m-%d").date()
    rows = VehicleExpense.query.filter(VehicleExpense.exp_date >= sdate, VehicleExpense.exp_date <= edate)\
        .order_by(VehicleExpense.exp_date).all()
    wb = Workbook(); ws = wb.active; ws.title = "Spese Veicoli"
    ws.append(["Data","Veicolo","Tipo","Importo","Ricevuta"])
    for c in range(1,6): ws.cell(row=1, column=c).font = Font(bold=True)
    tot=0.0
    for r in rows:
        tot += float(r.amount or 0.0)
        ws.append([r.exp_date.strftime("%Y-%m-%d"), r.vehicle.plate if r.vehicle else "", r.exp_type, float(r.amount or 0.0), r.receipt_path or ""])
    ws.append(["","","Totale", tot,""]); ws.cell(row=ws.max_row, column=3).font = Font(bold=True); ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"spese_veicoli_{sdate.strftime('%Y%m%d')}_{edate.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/export/spese-cantiere")
@login_required
def export_spese_cantiere():
    if not is_admin(): return redirect(url_for("dashboard"))
    start = request.args.get("start") or (date.today() - timedelta(days=7)).isoformat()
    end = request.args.get("end") or date.today().isoformat()
    sdate = datetime.strptime(start,"%Y-%m-%d").date()
    edate = datetime.strptime(end,"%Y-%m-%d").date()
    rows = SiteExpense.query.filter(SiteExpense.exp_date >= sdate, SiteExpense.exp_date <= edate)\
        .order_by(SiteExpense.exp_date).all()
    wb = Workbook(); ws = wb.active; ws.title = "Spese Cantiere"
    ws.append(["Data","Cantiere","Tipo","Pagamento","Importo","Ricevuta"])
    for c in range(1,7): ws.cell(row=1, column=c).font = Font(bold=True)
    tot=0.0
    for r in rows:
        tot += float(r.amount or 0.0)
        ws.append([r.exp_date.strftime("%Y-%m-%d"), r.site.name if r.site else "", r.exp_type, r.payment_type or "", float(r.amount or 0.0), r.receipt_path or ""])
    ws.append(["","","","Totale", tot,""]); ws.cell(row=ws.max_row, column=4).font = Font(bold=True); ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"spese_cantiere_{sdate.strftime('%Y%m%d')}_{edate.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ------------------ Template helpers ------------------
@app.context_processor
def inject_now():
    return {"now": datetime.now(), "timedelta": timedelta, "date": date}

@app.context_processor
def url_for_positional_adapter():
    def url_for_fix(endpoint, *args, **kwargs):
        if args:
            try:
                if endpoint == "delete_clients_sites" and len(args) == 2:
                    kind, oid = args
                    return url_for(endpoint, kind=kind, oid=oid)
                if endpoint == "delete_catalogo" and len(args) == 1:
                    (aid,) = args
                    return url_for(endpoint, aid=aid)
                if endpoint == "delete_user" and len(args) == 1:
                    (uid,) = args
                    return url_for(endpoint, uid=uid)
                if endpoint == "delete_vehicle" and len(args) == 1:
                    (vid,) = args
                    return url_for(endpoint, vid=vid)
                if endpoint == "equipment_delete" and len(args) == 1:
                    (eid,) = args
                    return url_for(endpoint, eid=eid)
            except Exception:
                pass
        return url_for(endpoint, **kwargs)
    return dict(url_for=url_for_fix)

# ------------------ Run ------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
