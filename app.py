import os
import io
from datetime import datetime, date, timedelta
from collections import defaultdict

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, login_user, login_required, logout_user,
    current_user, UserMixin
)
from sqlalchemy import func, or_
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.styles import Font

# ------------------ App config ------------------
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("APP_SECRET_KEY", "dev-secret")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///data.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

BASE_STATIC = os.path.join(os.path.dirname(__file__), "static")
UPLOAD_DIR = os.path.join(BASE_STATIC, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif", "pdf"}
def allowed_file(name: str) -> bool:
    return "." in name and name.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"

# ------------------ Models ------------------
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), default="capo")  # "admin" or "capo"
    full_name = db.Column(db.String(120))

    def set_password(self, p):
        self.password_hash = generate_password_hash(p)

    def check_password(self, p):
        return check_password_hash(self.password_hash, p)

class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)

class Site(db.Model):  # cantiere
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False)
    client = db.relationship("Client")

class ActivityCatalog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(40), nullable=False, unique=True)
    description = db.Column(db.String(255), nullable=False)
    unit = db.Column(db.String(20), nullable=False)  # unità di misura
    unit_price = db.Column(db.Float, nullable=False, default=0.0)

class ClientActivity(db.Model):
    """Attività assegnata ad un cantiere (tipologia + quantità iniziale)"""
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey("site.id"), nullable=False)
    activity_id = db.Column(db.Integer, db.ForeignKey("activity_catalog.id"), nullable=False)
    initial_qty = db.Column(db.Float, default=0.0)

    site = db.relationship("Site")
    activity = db.relationship("ActivityCatalog")

class Vehicle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    plate = db.Column(db.String(20), unique=True, nullable=False)
    description = db.Column(db.String(120))
    site_id = db.Column(db.Integer, db.ForeignKey("site.id"))
    site = db.relationship("Site")

class Equipment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(40), unique=True, nullable=False)
    description = db.Column(db.String(255), nullable=False)
    qty = db.Column(db.Integer, default=1)
    status = db.Column(db.String(20), default="disponibile")  # disponibile, occupato, rotto, manutenzione
    site_id = db.Column(db.Integer, db.ForeignKey("site.id"))
    site = db.relationship("Site")

class Assignment(db.Model):
    """Associazione capocantiere -> cantiere"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    site_id = db.Column(db.Integer, db.ForeignKey("site.id"), nullable=False)
    user = db.relationship("User")
    site = db.relationship("Site")

class ActivityEntry(db.Model):
    """Rilevazione giornaliera quantità lavorata"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)  # chi inserisce
    site_id = db.Column(db.Integer, db.ForeignKey("site.id"), nullable=False)
    client_activity_id = db.Column(db.Integer, db.ForeignKey("client_activity.id"), nullable=False)
    work_date = db.Column(db.Date, nullable=False, default=date.today)
    qty = db.Column(db.Float, nullable=False, default=0.0)
    note = db.Column(db.String(255))

    user = db.relationship("User")
    site = db.relationship("Site")
    client_activity = db.relationship("ClientActivity")

class SiteExpense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey("site.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    exp_type = db.Column(db.String(120), nullable=False)  # testo libero
    amount = db.Column(db.Float, nullable=False)
    payment_type = db.Column(db.String(40))  # contanti/pos/bonifico/altro
    exp_date = db.Column(db.Date, default=date.today, nullable=False)
    receipt_path = db.Column(db.String(255))

    site = db.relationship("Site")
    user = db.relationship("User")

class VehicleExpense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey("vehicle.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    exp_type = db.Column(db.String(40), nullable=False)  # CARBURANTE/PEDAGGIO/PARCHEGGIO
    amount = db.Column(db.Float, nullable=False)
    exp_date = db.Column(db.Date, default=date.today, nullable=False)
    receipt_path = db.Column(db.String(255))
    vehicle = db.relationship("Vehicle")
    user = db.relationship("User")

class Alert(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    message = db.Column(db.String(255), nullable=False)
    is_read = db.Column(db.Boolean, default=False)

# ------------------ Login ------------------
@login_manager.user_loader
def load_user(uid):
    # rimuove la warning legacy
    return db.session.get(User, int(uid))

# ------------------ Helpers ------------------
def is_admin():
    return current_user.is_authenticated and current_user.role == "admin"

def save_upload(file_storage, prefix):
    if not file_storage or not file_storage.filename:
        return None
    fname = secure_filename(file_storage.filename)
    if not allowed_file(fname):
        return None
    out = f"{prefix}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{fname}"
    path = os.path.join(UPLOAD_DIR, out)
    file_storage.save(path)
    return f"uploads/{out}"

def user_site_ids(user_id: int):
    """Ritorna gli ID dei cantieri assegnati all'utente (Assignment)."""
    return [a.site_id for a in Assignment.query.filter_by(user_id=user_id).all()]

# ------------------ Init DB ------------------
def ensure_demo():
    db.create_all()
    if not User.query.filter_by(username="admin").first():
        admin = User(username="admin", role="admin", full_name="Amministratore")
        admin.set_password("admin")
        db.session.add(admin)
    if not Client.query.first():
        c = Client(name="Cliente Alpha")
        db.session.add(c)
        s1 = Site(name="Cantiere Nord", client=c)
        s2 = Site(name="Cantiere Sud", client=c)
        db.session.add_all([s1, s2])
    if not ActivityCatalog.query.first():
        db.session.add_all([
            ActivityCatalog(code="SCAVO", description="Scavo generale", unit="m3", unit_price=25),
            ActivityCatalog(code="MUR", description="Muratura", unit="mq", unit_price=45),
        ])
    if not User.query.filter_by(username="capo").first():
        u = User(username="capo", role="capo", full_name="Mario Rossi")
        u.set_password("capo")
        db.session.add(u)
    db.session.commit()

with app.app_context():
    ensure_demo()

# ------------------ Routes (auth) ------------------
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","")
        u = User.query.filter_by(username=username).first()
        if u and u.check_password(password):
            login_user(u)
            flash("Accesso eseguito", "success")
            return redirect(url_for("dashboard"))
        flash("Credenziali errate", "danger")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logout effettuato", "info")
    return redirect(url_for("index"))

@app.route("/dashboard")
@login_required
def dashboard():
    if is_admin():
        unread = Alert.query.filter_by(is_read=False).count()
        return render_template("dashboard_admin.html", unread=unread)
    # Capocantiere (operatore)
    my_sites = (
        db.session.query(Site)
        .join(Assignment, Assignment.site_id == Site.id)
        .filter(Assignment.user_id == current_user.id)
        .all()
    )
    return render_template("dashboard_capo.html", sites=my_sites)

# ------------------ Admin: anagrafiche ------------------
@app.route("/admin/clients-sites", methods=["GET","POST"])
@login_required
def clients_sites():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        kind = request.form.get("kind")
        name = request.form.get("name","").strip()
        if not name:
            flash("Nome obbligatorio", "warning"); return redirect(url_for("clients_sites"))
        if kind == "client":
            db.session.add(Client(name=name))
        elif kind == "site":
            client_id = request.form.get("client_id", type=int)
            if not client_id: 
                flash("Cliente obbligatorio per il cantiere", "warning"); return redirect(url_for("clients_sites"))
            db.session.add(Site(name=name, client_id=client_id))
        db.session.commit()
        flash("Salvato", "success")
        return redirect(url_for("clients_sites"))
    return render_template("clients_sites.html",
                           clients=Client.query.order_by(Client.name).all(),
                           sites=Site.query.order_by(Site.name).all())

@app.route("/admin/clients-sites/<string:kind>/<int:oid>/delete", methods=["POST"])
@login_required
def delete_clients_sites(kind, oid):
    if not is_admin(): return redirect(url_for("dashboard"))
    model = Client if kind=="client" else Site
    obj = model.query.get_or_404(oid)
    db.session.delete(obj); db.session.commit()
    flash("Eliminato", "success")
    return redirect(url_for("clients_sites"))

# ------------------ Admin: Catalogo attività e assegnazioni ------------------
@app.route("/admin/catalogo", methods=["GET","POST"])
@login_required
def catalogo():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        code = request.form.get("code","").strip().upper()
        description = request.form.get("description","").strip()
        unit = request.form.get("unit","").strip()
        unit_price = request.form.get("unit_price", type=float) or 0.0
        if not code or not description or not unit:
            flash("Campi mancanti", "warning")
        else:
            if ActivityCatalog.query.filter_by(code=code).first():
                flash("Codice già presente", "danger")
            else:
                db.session.add(ActivityCatalog(code=code, description=description, unit=unit, unit_price=unit_price))
                db.session.commit()
                flash("Attività inserita", "success")
        return redirect(url_for("catalogo"))
    return render_template("catalogo.html", items=ActivityCatalog.query.order_by(ActivityCatalog.code).all())

@app.route("/admin/catalogo/<int:aid>/delete", methods=["POST"])
@login_required
def delete_catalogo(aid):
    if not is_admin(): return redirect(url_for("dashboard"))
    a = ActivityCatalog.query.get_or_404(aid)
    db.session.delete(a); db.session.commit()
    flash("Eliminato", "success")
    return redirect(url_for("catalogo"))

@app.route("/admin/assign-activity", methods=["POST"])
@login_required
def assign_activity():
    if not is_admin(): return redirect(url_for("dashboard"))
    site_id = request.form.get("site_id", type=int)
    activity_id = request.form.get("activity_id", type=int)
    initial_qty = request.form.get("initial_qty", type=float) or 0.0
    if not site_id or not activity_id:
        flash("Seleziona cantiere e attività", "warning")
    else:
        db.session.add(ClientActivity(site_id=site_id, activity_id=activity_id, initial_qty=initial_qty))
        db.session.commit()
        flash("Assegnazione creata", "success")
    return redirect(url_for("catalogo"))

# ------------------ Admin: Utenti e assegnazioni cantiere ------------------
@app.route("/admin/users", methods=["GET","POST"])
@login_required
def users():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","")
        role = request.form.get("role","capo")
        full_name = request.form.get("full_name","").strip()
        if not username or not password:
            flash("Username e password obbligatori", "warning")
        else:
            if User.query.filter_by(username=username).first():
                flash("Username già in uso", "danger")
            else:
                u = User(username=username, role=role, full_name=full_name)
                u.set_password(password)
                db.session.add(u); db.session.commit()
                flash("Utente creato", "success")
        return redirect(url_for("users"))
    return render_template("users.html",
                           users=User.query.order_by(User.username).all(),
                           sites=Site.query.order_by(Site.name).all(),
                           assignments=Assignment.query.all())

@app.route("/admin/users/<int:uid>/delete", methods=["POST"])
@login_required
def delete_user(uid):
    if not is_admin(): return redirect(url_for("dashboard"))
    if current_user.id == uid:
        flash("Non puoi eliminare te stesso", "warning")
        return redirect(url_for("users"))
    u = User.query.get_or_404(uid)
    db.session.delete(u); db.session.commit()
    flash("Utente eliminato", "success")
    return redirect(url_for("users"))

@app.route("/admin/assign-site", methods=["POST"])
@login_required
def assign_site():
    if not is_admin(): return redirect(url_for("dashboard"))
    user_id = request.form.get("user_id", type=int)
    site_id = request.form.get("site_id", type=int)
    if user_id and site_id:
        if not Assignment.query.filter_by(user_id=user_id, site_id=site_id).first():
            db.session.add(Assignment(user_id=user_id, site_id=site_id))
            db.session.commit()
            flash("Cantiere assegnato", "success")
    return redirect(url_for("users"))

@app.route("/admin/assign-site/<int:aid>/delete", methods=["POST"])
@login_required
def delete_assignment(aid):
    if not is_admin(): return redirect(url_for("dashboard"))
    a = Assignment.query.get_or_404(aid)
    db.session.delete(a); db.session.commit()
    flash("Assegnazione rimossa", "success")
    return redirect(url_for("users"))

# ------------------ Admin: Veicoli e Attrezzature ------------------
@app.route("/admin/vehicles", methods=["GET","POST"])
@login_required
def vehicles():
    if not is_admin(): return redirect(url_for("dashboard"))
    if request.method == "POST":
        plate = request.form.get("plate","").upper().strip()
        description = request.form.get("description","").strip()
        site_id = request.form.get("site_id", type=int)
        if not plate:
            flash("Targa obbligatoria", "warning")
        else:
            db.session.add(Vehicle(plate=plate, description=description, site_id=site_id))
            db.session.commit()
            flash("Veicolo inserito", "success")
        return redirect(url_for("vehicles"))
    return render_template("vehicles.html",
                           vehicles=Vehicle.query.order_by(Vehicle.plate).all(),
                           sites=Site.query.order_by(Site.name).all())

@app.route("/admin/vehicles/<int:vid>/delete", methods=["POST"])
@login_required
def delete_vehicle(vid):
    if not is_admin(): return redirect(url_for("dashboard"))
    v = Vehicle.query.get_or_404(vid)
    db.session.delete(v); db.session.commit()
    flash("Veicolo eliminato", "success")
    return redirect(url_for("vehicles"))

# ================== ATTREZZATURE ==================
@app.route("/admin/equipment", methods=["GET","POST"])
@login_required
def equipment():
    """
    GET:
      - Admin: vede tutto.
      - Capo: vede solo attrezzature dei propri cantieri.
    POST (crea/assegna/elimina): solo Admin.
    """
    # --- POST: solo admin crea/assegna/elimina ---
    if request.method == "POST":
        if not is_admin():
            flash("Non autorizzato", "danger")
            return redirect(url_for("equipment"))

        # elimina con hidden delete_eid (compatibile con template)
        delete_eid = request.form.get("delete_eid")
        if delete_eid:
            try:
                Equipment.query.filter_by(id=int(delete_eid)).delete()
                db.session.commit()
                flash("Attrezzatura eliminata", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Errore eliminazione: {e}", "danger")
            return redirect(url_for("equipment"))

        # crea/assegna
        code = (request.form.get("code") or "").strip().upper()
        description = (request.form.get("description") or "").strip()
        qty = request.form.get("qty", type=int) or 1
        site_id = request.form.get("site_id", type=int)
        status = request.form.get("status","disponibile")
        if not code or not description:
            flash("Codice e descrizione obbligatori", "warning")
        else:
            try:
                db.session.add(Equipment(code=code, description=description, qty=qty, site_id=site_id, status=status))
                db.session.commit()
                flash("Attrezzatura inserita/assegnata", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Errore inserimento: {e}", "danger")
        return redirect(url_for("equipment"))

    # --- GET: prepara dati per template ---
    alerts = Alert.query.order_by(Alert.created_at.desc()).limit(15).all()

    if is_admin():
        equipments = Equipment.query.order_by(Equipment.code).all()
        sites = Site.query.order_by(Site.name).all()
        my_sites = sites
    else:
        ids = user_site_ids(current_user.id)
        sites = Site.query.filter(Site.id.in_(ids)).order_by(Site.name).all() if ids else []
        my_sites = sites
        equipments = (
            Equipment.query.filter(Equipment.site_id.in_(ids)).order_by(Equipment.code).all()
            if ids else []
        )

    return render_template("equipment.html",
                           equipments=equipments,
                           sites=sites,
                           my_sites=my_sites,
                           alerts=alerts)

@app.route("/admin/equipment/<int:eid>/status", methods=["POST"])
@login_required
def equipment_status(eid):
    """
    Aggiorna lo stato dell'attrezzatura.
    - Admin: qualsiasi stato (disponibile/occupato/rotto/manutenzione).
    - Capo: SOLO 'rotto' o 'manutenzione' e SOLO su attrezzature dei propri cantieri.
    """
    allowed_all = ["disponibile", "occupato", "rotto", "manutenzione"]
    allowed_capo = ["rotto", "manutenzione"]

    new_status = (request.form.get("status") or "").strip().lower()
    if not new_status:
        flash("Stato non specificato", "warning")
        return redirect(url_for("equipment"))

    eq = db.session.get(Equipment, eid)
    if not eq:
        flash("Attrezzatura non trovata", "warning")
        return redirect(url_for("equipment"))

    if is_admin():
        if new_status not in allowed_all:
            flash("Stato non valido", "warning")
            return redirect(url_for("equipment"))
    else:
        # capo/operatore: vincoli
        if new_status not in allowed_capo:
            flash("Come operatore puoi impostare solo: rotto o manutenzione", "warning")
            return redirect(url_for("equipment"))
        my_ids = set(user_site_ids(current_user.id))
        if not eq.site_id or eq.site_id not in my_ids:
            flash("Non puoi modificare attrezzature non assegnate ai tuoi cantieri", "danger")
            return redirect(url_for("equipment"))

    try:
        eq.status = new_status
        db.session.commit()
        if not is_admin():
            db.session.add(Alert(message=f"Attrezzatura {eq.code} segnalata '{new_status}' dal capo {current_user.username}"))
            db.session.commit()
        flash("Stato aggiornato", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Errore aggiornamento: {e}", "danger")

    return redirect(url_for("equipment"))

@app.route("/admin/equipment/<int:eid>/delete", methods=["POST"])
@login_required
def delete_equipment(eid):
    if not is_admin(): return redirect(url_for("dashboard"))
    e = Equipment.query.get_or_404(eid)
    db.session.delete(e); db.session.commit()
    flash("Attrezzatura eliminata", "success")
    return redirect(url_for("equipment"))

# ------------------ Capocantiere: Attività giornaliere ------------------
@app.route("/capo/attivita", methods=["GET","POST"])
@login_required
def capo_attivita():
    if is_admin():  # admin può filtrare per cantiere e utente
        my_assignments = Assignment.query.all()
    else:
        my_assignments = Assignment.query.filter_by(user_id=current_user.id).all()

    if request.method == "POST":
        work_date = request.form.get("work_date") or date.today().isoformat()
        site_id = request.form.get("site_id", type=int)
        client_activity_id = request.form.get("client_activity_id", type=int)
        qty = request.form.get("qty", type=float) or 0.0
        note = request.form.get("note","").strip() or None
        if not site_id or not client_activity_id:
            flash("Seleziona cantiere e attività", "warning")
        else:
            entry = ActivityEntry(
                user_id=current_user.id,
                site_id=site_id,
                client_activity_id=client_activity_id,
                work_date=datetime.strptime(work_date,"%Y-%m-%d").date(),
                qty=qty,
                note=note
            )
            db.session.add(entry); db.session.commit()
            flash("Rilevazione salvata", "success")
        return redirect(url_for("capo_attivita"))

    # attività disponibili per i cantieri assegnati
    site_ids = [a.site_id for a in my_assignments]
    assigned_ca = ClientActivity.query.filter(ClientActivity.site_id.in_(site_ids)).all() if site_ids else []
    entries = ActivityEntry.query.order_by(ActivityEntry.work_date.desc(), ActivityEntry.id.desc())
    if not is_admin():
        entries = entries.filter(ActivityEntry.user_id == current_user.id)
    entries = entries.limit(200).all()
    return render_template("capo_attivita.html", assignments=my_assignments, assigned_ca=assigned_ca, entries=entries)

@app.route("/capo/attivita/<int:eid>/delete", methods=["POST"])
@login_required
def delete_entry(eid):
    e = ActivityEntry.query.get_or_404(eid)
    if not is_admin() and e.user_id != current_user.id:
        flash("Operazione non consentita", "danger"); return redirect(url_for("capo_attivita"))
    db.session.delete(e); db.session.commit()
    flash("Eliminato", "success")
    return redirect(url_for("capo_attivita"))

# ------------------ Spese cantiere (capo) ------------------
@app.route("/capo/spese", methods=["GET","POST"])
@login_required
def capo_spese():
    # elenco cantieri del capo
    assignments = Assignment.query.filter_by(user_id=current_user.id).all() if not is_admin() else Assignment.query.all()
    if request.method == "POST":
        site_id = request.form.get("site_id", type=int)
        exp_type = request.form.get("exp_type","").strip()
        amount = request.form.get("amount", type=float) or 0.0
        payment_type = request.form.get("payment_type","").strip() or None
        exp_date = request.form.get("exp_date") or date.today().isoformat()
        receipt = save_upload(request.files.get("receipt"), "SITEEXP")
        if not site_id or not exp_type or not amount:
            flash("Campi obbligatori mancanti", "warning")
        else:
            db.session.add(SiteExpense(
                site_id=site_id, user_id=current_user.id, exp_type=exp_type,
                amount=amount, payment_type=payment_type,
                exp_date=datetime.strptime(exp_date,"%Y-%m-%d").date(),
                receipt_path=receipt
            ))
            db.session.commit()
            flash("Spesa registrata", "success")
        return redirect(url_for("capo_spese"))
    # lista
    q = SiteExpense.query.order_by(SiteExpense.exp_date.desc(), SiteExpense.id.desc())
    if not is_admin():
        q = q.filter(SiteExpense.user_id == current_user.id)
    rows = q.limit(200).all()
    return render_template("capo_spese.html", assignments=assignments, rows=rows)

# ------------------ Spese veicoli (capo) ------------------
@app.route("/capo/spese-veicoli", methods=["GET","POST"])
@login_required
def capo_spese_veicoli():
    vehicles = Vehicle.query.order_by(Vehicle.plate).all()
    if request.method == "POST":
        vehicle_id = request.form.get("vehicle_id", type=int)
        exp_type = request.form.get("exp_type","").strip().upper()
        amount = request.form.get("amount", type=float) or 0.0
        exp_date = request.form.get("exp_date") or date.today().isoformat()
        receipt = save_upload(request.files.get("receipt"), "VEH")
        if not vehicle_id or not exp_type or not amount:
            flash("Campi obbligatori mancanti", "warning")
        else:
            db.session.add(VehicleExpense(
                vehicle_id=vehicle_id, user_id=current_user.id,
                exp_type=exp_type, amount=amount,
                exp_date=datetime.strptime(exp_date,"%Y-%m-%d").date(),
                receipt_path=receipt
            ))
            db.session.commit()
            flash("Spesa veicolo registrata", "success")
        return redirect(url_for("capo_spese_veicoli"))
    q = VehicleExpense.query.order_by(VehicleExpense.exp_date.desc(), VehicleExpense.id.desc())
    if not is_admin():
        q = q.filter(VehicleExpense.user_id == current_user.id)
    rows = q.limit(200).all()
    return render_template("capo_spese_veicoli.html", vehicles=vehicles, rows=rows)

# ------------------ Capo: attrezzature status (alert admin) ------------------
@app.route("/capo/equipment/<int:eid>/status", methods=["POST"])
@login_required
def capo_equipment_status(eid):
    e = Equipment.query.get_or_404(eid)
    new_status = request.form.get("status","manutenzione")
    # vincoli lato funzione "capo": solo rotto/manutenzione e solo sui propri cantieri
    if new_status not in ("rotto","manutenzione"):
        flash("Come operatore puoi impostare solo: rotto o manutenzione", "warning")
        return redirect(url_for("dashboard"))
    if e.site_id not in set(user_site_ids(current_user.id)):
        flash("Attrezzatura non assegnata ad un tuo cantiere", "danger")
        return redirect(url_for("dashboard"))
    e.status = new_status
    db.session.add(Alert(message=f"Attrezzatura {e.code} segnalata '{new_status}' dal capo {current_user.username}"))
    db.session.commit()
    flash("Segnalazione inviata", "success")
    return redirect(url_for("dashboard"))

# ------------------ Report ed export ------------------
@app.route("/admin/report", methods=["GET"])
@login_required
def admin_report():
    if not is_admin(): return redirect(url_for("dashboard"))
    # qualche numero
    today = date.today()
    start = datetime(today.year, today.month, 1).date()
    acts = db.session.query(ActivityEntry.work_date,
                            func.sum(ActivityEntry.qty * ActivityCatalog.unit_price).label("val"))
    acts = acts.join(ClientActivity, ClientActivity.id == ActivityEntry.client_activity_id)\
               .join(ActivityCatalog, ActivityCatalog.id == ClientActivity.activity_id)\
               .filter(ActivityEntry.work_date >= start)\
               .group_by(ActivityEntry.work_date)\
               .order_by(ActivityEntry.work_date).all()
    total_val = sum([float(a.val or 0) for a in acts])
    veh_cost = db.session.query(func.sum(VehicleExpense.amount)).scalar() or 0.0
    site_cost = db.session.query(func.sum(SiteExpense.amount)).scalar() or 0.0

    # default per i campi data nel template (evita usare timedelta in jinja)
    start_default = (date.today() - timedelta(days=7)).isoformat()
    end_default = date.today().isoformat()

    return render_template(
        "admin_report.html",
        total_val=total_val,
        veh_cost=veh_cost,
        site_cost=site_cost,
        start_default=start_default,
        end_default=end_default
    )

@app.route("/admin/export/attivita")
@login_required
def export_attivita():
    if not is_admin(): return redirect(url_for("dashboard"))
    start = request.args.get("start") or (date.today() - timedelta(days=7)).isoformat()
    end = request.args.get("end") or date.today().isoformat()
    sdate = datetime.strptime(start,"%Y-%m-%d").date()
    edate = datetime.strptime(end,"%Y-%m-%d").date()

    rows = db.session.query(
        ActivityEntry.work_date, Site.name, Client.name,
        ActivityCatalog.code, ActivityCatalog.description, ActivityCatalog.unit,
        ActivityEntry.qty, ActivityCatalog.unit_price,
        (ActivityEntry.qty * ActivityCatalog.unit_price).label("value")
    ).join(Site, Site.id == ActivityEntry.site_id)\
     .join(Client, Client.id == Site.client_id)\
     .join(ClientActivity, ClientActivity.id == ActivityEntry.client_activity_id)\
     .join(ActivityCatalog, ActivityCatalog.id == ClientActivity.activity_id)\
     .filter(ActivityEntry.work_date >= sdate, ActivityEntry.work_date <= edate)\
     .order_by(ActivityEntry.work_date).all()

    wb = Workbook(); ws = wb.active; ws.title = "Attività"
    ws.append(["Data","Cantiere","Cliente","Codice","Descrizione","UM","Q.tà","Prezzo Unit.","Valore (€)"])
    for c in range(1,10): ws.cell(row=1, column=c).font = Font(bold=True)
    total = 0.0
    for r in rows:
        total += float(r.value or 0.0)
        ws.append([r.work_date.strftime("%Y-%m-%d"), r.name, r[2], r.code, r.description, r.unit, float(r.qty or 0), float(r.unit_price or 0), float(r.value or 0)])
    ws.append(["","","","","","","","Totale", total]); ws.cell(row=ws.max_row, column=8).font = Font(bold=True); ws.cell(row=ws.max_row, column=9).font = Font(bold=True)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"attivita_{sdate.strftime('%Y%m%d')}_{edate.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/export/spese-veicoli")
@login_required
def export_spese_veicoli():
    if not is_admin(): return redirect(url_for("dashboard"))
    start = request.args.get("start") or (date.today() - timedelta(days=7)).isoformat()
    end = request.args.get("end") or date.today().isoformat()
    sdate = datetime.strptime(start,"%Y-%m-%d").date()
    edate = datetime.strptime(end,"%Y-%m-%d").date()
    rows = VehicleExpense.query.filter(VehicleExpense.exp_date >= sdate, VehicleExpense.exp_date <= edate)\
        .order_by(VehicleExpense.exp_date).all()
    wb = Workbook(); ws = wb.active; ws.title = "Spese Veicoli"
    ws.append(["Data","Veicolo","Tipo","Importo","Ricevuta"])
    for c in range(1,6): ws.cell(row=1, column=c).font = Font(bold=True)
    tot=0.0
    for r in rows:
        tot += float(r.amount or 0.0)
        ws.append([r.exp_date.strftime("%Y-%m-%d"), r.vehicle.plate, r.exp_type, float(r.amount or 0.0), r.receipt_path or ""])
    ws.append(["","","Totale", tot,""]); ws.cell(row=ws.max_row, column=3).font = Font(bold=True); ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"spese_veicoli_{sdate.strftime('%Y%m%d')}_{edate.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/export/spese-cantiere")
@login_required
def export_spese_cantiere():
    if not is_admin(): return redirect(url_for("dashboard"))
    start = request.args.get("start") or (date.today() - timedelta(days=7)).isoformat()
    end = request.args.get("end") or date.today().isoformat()
    sdate = datetime.strptime(start,"%Y-%m-%d").date()
    edate = datetime.strptime(end,"%Y-%m-%d").date()
    rows = SiteExpense.query.filter(SiteExpense.exp_date >= sdate, SiteExpense.exp_date <= edate)\
        .order_by(SiteExpense.exp_date).all()
    wb = Workbook(); ws = wb.active; ws.title = "Spese Cantiere"
    ws.append(["Data","Cantiere","Tipo","Pagamento","Importo","Ricevuta"])
    for c in range(1,7): ws.cell(row=1, column=c).font = Font(bold=True)
    tot=0.0
    for r in rows:
        tot += float(r.amount or 0.0)
        ws.append([r.exp_date.strftime("%Y-%m-%d"), r.site.name, r.exp_type, r.payment_type or "", float(r.amount or 0.0), r.receipt_path or ""])
    ws.append(["","","","Totale", tot,""]); ws.cell(row=ws.max_row, column=4).font = Font(bold=True); ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"spese_cantiere_{sdate.strftime('%Y%m%d')}_{edate.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ------------------ Template helpers ------------------
@app.context_processor
def inject_now():
    # Rendo disponibili oggetti utili nei template
    return {
        "now": datetime.now(),
        "timedelta": timedelta,  # se qualche template lo usa direttamente
        "date": date
    }

# Adattatore per supportare chiamate url_for con argomenti POSIZIONALI dai template esistenti
@app.context_processor
def url_for_positional_adapter():
    def url_for_fix(endpoint, *args, **kwargs):
        # Se arrivano argomenti posizionali dai template, li mappo ai nomi corretti
        if args:
            try:
                if endpoint == "delete_clients_sites" and len(args) == 2:
                    kind, oid = args
                    return url_for(endpoint, kind=kind, oid=oid)
                if endpoint == "delete_catalogo" and len(args) == 1:
                    (aid,) = args
                    return url_for(endpoint, aid=aid)
                if endpoint == "delete_user" and len(args) == 1:
                    (uid,) = args
                    return url_for(endpoint, uid=uid)
                if endpoint == "delete_vehicle" and len(args) == 1:
                    (vid,) = args
                    return url_for(endpoint, vid=vid)
            except Exception:
                # fallback: lascio passare a url_for, ma senza args posizionali
                pass
        return url_for(endpoint, **kwargs)
    return dict(url_for=url_for_fix)

# ------------------ Run ------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
